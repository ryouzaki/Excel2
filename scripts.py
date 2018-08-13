import argparse
from openpyxl import *
from openpyxl.cell import Cell
from openpyxl.styles import Border, Side, GradientFill, PatternFill, Alignment, Font
import re
import os
import datetime
import requests
import json
from copy import copy
import xml.etree.ElementTree as etree
import pymongo
from pymongo import MongoClient
from openpyxl.utils import column_index_from_string


def parse_plan(filename):
    plan = load_workbook(filename=filename, data_only=True)
    mediaplan_sheet = plan['MediaPlan']
    fields_row = get_fields_row(mediaplan_sheet)
    for row in mediaplan_sheet.iter_rows(min_col=2,max_col=mediaplan_sheet.max_column):
        is_placement_row = False
        for cell in row:
            if re.search('\w\w_\w\w_\d\d\d', str(cell.value)):
                is_placement_row = True
                placement_id = str(cell.value)
            if is_placement_row and mediaplan_sheet.cell(row=fields_row,column=column_index_from_string(cell.column)).value:
                updateDatabase(placement_id=placement_id, field=mediaplan_sheet.cell(row=fields_row,column=column_index_from_string(cell.column)).value, value=cell.value)
    return


def make_placement_dict(mediaplan_sheet, placement_row):
    placement_dict = dict()
    weeknumber_set = set()
    fields_row = get_fields_row(mediaplan_sheet)
    for row in mediaplan_sheet.iter_rows(min_row=placement_row,max_row=placement_row):
        for cell in row:
            if mediaplan_sheet.cell(row=fields_row, column=column_index_from_string(cell.column)).value is not None:
                if re.search('\d{1,2}', str(mediaplan_sheet.cell(row=fields_row, column=column_index_from_string(cell.column)).value)):
                    if cell.value == 1:
                        weeknumber_set.add(int(mediaplan_sheet.cell(row=fields_row, column=column_index_from_string(cell.column)).value))
                else:
                    placement_dict.update({mediaplan_sheet.cell(row=fields_row, column=column_index_from_string(cell.column)).value:cell.value})
    placement_dict.update({'plan_weeks':list(weeknumber_set)})
    return placement_dict


def parse_amnet(week):
    week_number = int(week)
    placement_id = None
    report = load_workbook(filename=os.getcwd()+'\Reports\\Amnet\\Amnet'+ str(week)+ '.xlsx', data_only=True)
    week_report_sheet = report[str(week_number)]
    for row in week_report_sheet.iter_rows(min_col=1,max_col=1):
        for cell in row:
            if re.search('\w\w_\w\w_\d\d\d', str(cell.value)):
                placement_id = str(cell.value)
                outfile = open(str(os.getcwd()) + '\\MP\\JSON\\' + placement_id + '.json', 'r+')
                placement_dict = json.load(outfile)
                if type(placement_dict.get('postclick')) is list:
                    new_week_list = list(value for value in placement_dict.get('postclick') if value['weeknumber'] != week_number)
                    print (new_week_list)
                    this_week_list = {'weeknumber': week_number, 'fact_budget': None, 'fact_impressions': None, 'fact_views': None,'fact_clicks': None,'fact_reach': None}
                    new_week_list.append(this_week_list)
                    placement_dict.update({'postclick': new_week_list})
                    outfile.seek(0)
                    outfile.truncate()
                    json.dump(placement_dict, outfile)
                else:
                    week_list = [{'weeknumber': week_number, 'fact_budget': None, 'fact_impressions': None, 'fact_views': None,'fact_clicks': None,'fact_reach': None}]
                    placement_dict.update({'postclick':week_list})
                    outfile.seek(0)
                    outfile.truncate()
                    json.dump(placement_dict, outfile)
            if placement_id:
                outfile = open(str(os.getcwd()) + '\\MP\\JSON\\' + placement_id + '.json', 'r+')
                placement_dict = json.load(outfile)
                for value in placement_dict.get('postclick'):
                    if int(value.get('weeknumber')) == int(week):
                        if str(cell.value) in value.keys():
                            outfile = open(str(os.getcwd()) + '\\MP\\JSON\\' + placement_id + '.json', 'r+')
                            value[str(cell.value)]=week_report_sheet.cell(row = cell.row, column = column_index_from_string(cell.column)+2).value
                            outfile.seek(0)
                            outfile.truncate()
                            json.dump(placement_dict, outfile)
    return


def parse_iprospect(week):
    week_number = int(week)
    report = load_workbook(filename=os.getcwd()+'\Reports\\iProspect\\iProspect'+str(week_number)+'.xlsx', data_only=True)
    week_report_sheet = report[str(week_number)]
    fields_row = 1
    for row in week_report_sheet.iter_rows(min_row=1, max_col=20, max_row=100):
        for cell in row:
            if cell.value == 'fact_impressions':
                fields_row = cell.row
    for row in week_report_sheet.iter_rows(min_col=1, max_col=1):
        for cell in row:
            if re.search('\w\w_\w\w_\d\d\d',str(cell.value)):
                placement_id = str(cell.value)
                print (placement_id)
                outfile = open(str(os.getcwd()) + '\\MP\\JSON\\' + placement_id + '.json', 'r+')
                placement_dict = json.load(outfile)
                if type(placement_dict.get('postclick')) is list:
                    new_week_list = list(value for value in placement_dict.get('postclick') if value['weeknumber'] != week)
                    this_week_list = {'weeknumber': week_number, 'fact_budget': None, 'fact_impressions': None, 'fact_views': None,'fact_clicks': None,'fact_reach': None}
                    new_week_list.append(this_week_list)
                    placement_dict.update({'postclick': new_week_list})
                    outfile.seek(0)
                    outfile.truncate()
                    json.dump(placement_dict, outfile)
                else:
                    week_list = [{'weeknumber': week_number, 'fact_budget': None, 'fact_impressions': None, 'fact_views': None,'fact_clicks': None,'fact_reach': None}]
                    placement_dict.update({'postclick':week_list})
                    outfile.seek(0)
                    outfile.truncate()
                    json.dump(placement_dict, outfile)
                for row1 in week_report_sheet.iter_rows(min_row=cell.row, max_row=cell.row):
                    for cell1 in row1:
                        for value in placement_dict.get('postclick'):
                            if int(value.get('weeknumber')) == int(week):
                                if str(week_report_sheet.cell(row = fields_row, column = column_index_from_string(cell1.column)).value) in value.keys():
                                    value[str(week_report_sheet.cell(row = fields_row, column = column_index_from_string(cell1.column)).value)]=cell1.value
                                    outfile.seek(0)
                                    outfile.truncate()
                                    json.dump(placement_dict, outfile)
    return

def create_report():
    #список недель для формирования необходимых листов
    report = load_workbook(filename=os.getcwd()+'\Reports\\Client\\'+ 'Template.xlsm', read_only=False, keep_vba=True)
    postclick_week_set = set()
    plan_week_set = set()
    for jsonplan in os.listdir(os.getcwd() + '\\MP\\JSON\\'):
        if jsonplan.endswith('.json'):
            with open(str(os.getcwd()) + '\\MP\\JSON\\' + jsonplan, 'r') as infile:
                placement_dict = json.load(infile)
                plan_week_set.update(placement_dict.get('plan_weeks'))
                for value in placement_dict['postclick']:
                    postclick_week_set.add(int(value['weeknumber']))
    for week in sorted(postclick_week_set):
        #начало переменных форматирования
        medium = Side(border_style="medium", color="000000")
        borders = Border (top=medium, left=medium, right=medium, bottom=medium)
        yellow_fill = GradientFill(stop=("ffff99", "ffff99"))
        text_rotation = Alignment(textRotation=90, horizontal='center', vertical='center')
        #конец переменных форматирования
        source = report.active
        target = report.copy_worksheet(source)
        fields_row=get_fields_row(target)
        first_stage_row= target.max_row+1

        for stage in ['Awareness','Consideration', 'Preference', 'Action', 'Loyalty']:
            if first_stage_row != target.max_row+1:
                target.cell(row=first_stage_row, column = 2).alignment = text_rotation
                target.merge_cells(start_row=first_stage_row, end_row=target.max_row, start_column=2, end_column=2)
                style_merged_cells(target,first_stage_row,target.max_row,2,2,border=borders, fill = yellow_fill)
                first_stage_row=target.max_row+1
            for category in ['OLV', 'Programmatic', 'Social Media', 'SEA']:
                category_flag = True
                for jsonplan in os.listdir(os.getcwd() + '\\MP\\JSON\\'):
                    if jsonplan.endswith('.json'):
                        with open(str(os.getcwd()) + '\\MP\\JSON\\' + jsonplan, 'r') as infile:
                            placement_dict = json.load(infile)
                            if placement_dict.get('stage') == stage and placement_dict.get('category') == category:
                                last_column = 1
                                last_row = target.max_row+1
                                if category_flag:
                                    target.cell(row=last_row, column=last_column+1, value=stage)
                                    target.merge_cells(start_row=last_row, start_column=last_column+2,end_row=last_row,end_column=target.max_column-1)
                                    target.cell(row=last_row, column=last_column+2, value = category)
                                    style_merged_cells(target, last_row, last_row, last_column+2, target.max_column-1, border=borders, fill = yellow_fill)
                                    category_flag = False
                                    last_row += 1
                                while target.cell(row=fields_row, column = last_column).value != "end":
                                    cell = get_value_from_placement_dict(target.cell(row=fields_row, column = last_column).value,placement_dict, week)
                                    target.cell(row=last_row,column=last_column, value=cell.value)
                                    target.cell(row=last_row,column=last_column).number_format = copy(target.cell(row=fields_row, column = last_column).number_format)
                                    target.cell(row=last_row,column=last_column).fill = copy(cell.fill)
                                    target.cell(row=last_row,column=last_column).alignment = copy(target.cell(row=fields_row, column=last_column).alignment)
                                    target.cell(row=last_row,column=last_column).border = copy(target.cell(row=fields_row, column=last_column).border)
                                    last_column+=1
        last_column = 1
        while target.cell(row=fields_row, column = last_column).value != "end":
            if re.search('\d{1,2}', str(target.cell(row=fields_row, column = last_column).value)) and (int(target.cell(row=fields_row, column = last_column).value) in plan_week_set or int(target.cell(row=fields_row, column = last_column).value) in postclick_week_set):
                target.cell(row=fields_row, column = last_column).font = Font(bold= True)
            total_formula = get_total(target.cell(row=fields_row, column = last_column), target)
            target.cell(row=last_row+1,column=last_column, value=total_formula)

            if total_formula != "":
                target.cell(row=last_row+1,column=last_column).border = Border(top=Side(style='medium'),bottom=Side(style='medium'),left=Side(style='thin'), right= Side(style='thin'))
            else:
                target.cell(row=last_row+1,column=last_column).border = Border(top=Side(style='medium'),bottom=Side(style='medium'))
            target.cell(row=last_row+1,column=last_column).font = Font(bold= True)
            target.cell(row=last_row+1,column=last_column).number_format = copy(target.cell(row=fields_row, column = last_column).number_format)
            target.cell(row=last_row+1,column=last_column).alignment = copy(target.cell(row=fields_row, column = last_column).alignment)
            last_column+=1
        d = str(datetime.datetime.now().year) + '-W' + str(week)
        target.title = datetime.datetime.strftime(datetime.datetime.strptime(d + '-1', "%Y-W%W-%w"),"%d-%b-%Y") + " -- " + datetime.datetime.strftime(datetime.datetime.strptime(d + '-1', "%Y-W%W-%w")+datetime.timedelta(days=6),"%d-%b-%Y")
    report.save(os.getcwd()+'\Reports\\Client\\'+ 'Report_week' +str(datetime.datetime.today().isocalendar()[1])+'.xlsm')
    return

def style_merged_cells(ws, min_row, max_row, min_col, max_col, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    first_cell = ws.cell(min_row,min_col)
    if alignment:
        ws.merge_cells(start_row=min_row,end_row=max_row,start_column=min_col,end_column=max_col)
        first_cell.alignment = alignment
    if font:
        first_cell.font = font
    for rows in ws.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row,max_row=max_row):
        for cell in rows:
            cell.border = cell.border + top
    for rows in ws.iter_rows(min_col=min_col, max_col=max_col,min_row=min_row, max_row=max_row):
        for cell in rows:
            cell.border = cell.border + bottom
    for columns in ws.iter_cols(min_col=min_col, max_col=max_col,min_row=min_row, max_row=max_row):
        for cell in columns:
            cell.border = cell.border + left
    for columns in ws.iter_cols(min_col=min_col, max_col=max_col,min_row=min_row, max_row=max_row):
        for cell in columns:
            cell.border = cell.border + right
            if fill:
                cell.fill=fill

def insert_fact_placement(target, placement_dict, week):
    fields_row=get_fields_row(target)
    print (placement_dict.get('fields'))
    count_of_creatives = len(placement_dict.get("creatives"))
    for columns in target.iter_cols(min_row=fields_row, max_row=fields_row):
        for cell in columns:
            if cell.value == "fields":
                return ""
            if cell.value in ("platform_site", "description", "format", "plan_impressions", "plan_reach", "plan_clicks", "plan_views", "plan_budget"):
                target.cell(row = target.max_row+1, column = column_index_from_string(cell.column)).value = placement_dict.get(cell.value)
                target.merge_cells(start_row=target.max_row+1, end_row= target.max_row+1+count_of_creatives, start_column=column_index_from_string(cell.column), end_column=column_index_from_string(cell.column))
            if cell.value == "plan_cpm":
                if placement_dict.get("plan_budget") is None or placement_dict.get("plan_impressions") is None or placement_dict.get("plan_budget") == "N/A" or placement_dict.get("plan_impressions") == "N/A":
                    return "N/A"
                else:
                    placement_dict.get("plan_budget")*1000/placement_dict.get("plan_impressions")
            if cell.value == "plan_cpt":
                if placement_dict.get("plan_budget") is None or placement_dict.get("plan_reach") is None or placement_dict.get("plan_budget") == "N/A" or placement_dict.get("plan_reach") == "N/A":
                    return "N/A"
                else:
                    return placement_dict.get("plan_budget")*1000/placement_dict.get("plan_reach")
            if cell.value == "plan_ctr":
                if placement_dict.get("plan_clicks") is None or placement_dict.get("plan_impressions") is None or placement_dict.get("plan_clicks") == "N/A" or placement_dict.get("plan_impressions") == "N/A":
                    return "N/A"
                else:
                    return placement_dict.get("plan_clicks")/placement_dict.get("plan_impressions")
            if cell.value == "plan_cpc":
                if placement_dict.get("plan_budget") is None or placement_dict.get("plan_clicks") is None or placement_dict.get("plan_budget") == "N/A" or placement_dict.get("plan_clicks") == "N/A":
                    return "N/A"
                else:
                    return placement_dict.get("plan_budget")/placement_dict.get("plan_clicks")
            if cell.value == "plan_vtr":
                if placement_dict.get("plan_impressions") is None or placement_dict.get("plan_impressions") is None or placement_dict.get("plan_impressions") == "N/A" or placement_dict.get("plan_views") == "N/A":
                    return "N/A"
                else:
                    return placement_dict.get("plan_views")/placement_dict.get("plan_impressions")

def get_fields_row(worksheet):
    fields_row=1
    while worksheet.cell(row=fields_row,column=1).value != "fields":
        fields_row+=1
    return fields_row


def get_adriver_value(adriver_id, week):
    if adriver_id:
        url = 'https://api.adriver.ru/login'
        headers = {'content-type': 'application/atom+xml','X-Auth-Login':'Adwatch2','X-Auth-Passwd':'RUG30qwm8zaV'}
        r = requests.get(url = url, headers=headers)
        tree = etree.fromstring(r.content)
        print(tree.findall('{http://adriver.ru/ns/restapi/atom}token')[0].text)
        d = str(datetime.datetime.now().year) + '-W' + str(week)
        stop_date = datetime.datetime.strftime(datetime.datetime.strptime(d + '-1', "%Y-W%W-%w")+datetime.timedelta(days=6),"%Y-%m-%d")
        print (stop_date)
        url = 'https://api.adriver.ru/stat/profiles/'+str(adriver_id)+'?start_date=2018-01-01&stop_date='+stop_date+'&granularity=daily'
        headers = {'content-type': 'application/atom+xml','X-Auth-Login':'130529','X-Auth-Passwd':'%5C%A8%F2i%A7%8CD%98%12%F8I'}
        r = requests.get(url = url, headers=headers)
        tree = etree.fromstring(r.content)
        return (sum(int(c.text) for c in tree.iter('{http://adriver.ru/ns/restapi/atom}exp')))


def get_value_from_placement_dict(attrib, placement_dict, week):
    black_solid_fill = PatternFill("solid", fgColor="3b3b3b")
    dark_grid_fill = PatternFill("darkGrid", fgColor="3b3b3b")
    wb = Workbook()
    ws = wb.active
    return_cell = Cell(ws)
    return_cell.value == ''
    if attrib == 'fields':
        return_cell.value == ''
    if re.search('\d{1,2}', str(attrib)):
        plan_weeks = placement_dict.get('plan_weeks')
        fact_weeks = list()
        for value in placement_dict['postclick']:
             fact_weeks.append(int(value['weeknumber']))
        if attrib in plan_weeks:
            return_cell.fill = dark_grid_fill
        if attrib in fact_weeks:
            return_cell.fill = dark_grid_fill
        if attrib in plan_weeks and attrib in fact_weeks:
            return_cell.fill = black_solid_fill
    if attrib in ('platform_site', 'description', 'format','plan_impressions', 'plan_reach', 'plan_clicks', 'plan_views','plan_budget'):
        return_cell.value = placement_dict.get(attrib)
    if attrib in ('fact_impressions', 'fact_clicks', "fact_reach", "fact_views", 'fact_budget'):
        for value in placement_dict['postclick']:
            if value['weeknumber'] == week:
                return_cell.value = value.get(attrib)
    if attrib == "plan_cpm":
        if placement_dict.get("plan_budget") is not None and placement_dict.get("plan_impressions"):
            return_cell.value = placement_dict.get("plan_budget")*1000/placement_dict.get("plan_impressions")
        else:
            return_cell.value = "N/A"
    if attrib == "plan_cpt":
        if placement_dict.get("plan_budget") is not None and placement_dict.get("plan_reach") is not None:
            return_cell.value = placement_dict.get("plan_budget")*1000/placement_dict.get("plan_reach")
        else:
            return_cell.value = "N/A"
    if attrib == "plan_ctr":
        if placement_dict.get("plan_clicks") is not None and placement_dict.get("plan_impressions") is not None:
            return_cell.value = placement_dict.get("plan_clicks")/placement_dict.get("plan_impressions")
        else:
            return_cell.value = "N/A"
    if attrib == "plan_cpc":
        if placement_dict.get("plan_budget") is not None and placement_dict.get("plan_clicks") is not None:
            return_cell.value = placement_dict.get("plan_budget")/placement_dict.get("plan_clicks")
        else:
            return_cell.value = "N/A"
    if attrib == "plan_vtr":
        if placement_dict.get("plan_views") != 'N/A' and placement_dict.get("plan_impressions") is not None:
            return_cell.value = placement_dict.get("plan_views")/placement_dict.get("plan_impressions")
        else:
            return_cell.value = "N/A"
    if attrib == "plan_cpv":
        if placement_dict.get("plan_budget") is not None and placement_dict.get("plan_views") != 'N/A':
            return_cell.value = placement_dict.get("plan_budget")/placement_dict.get("plan_views")
        else:
            return_cell.value = "N/A"
    if attrib == "fact_cpm":
        for value in placement_dict['postclick']:
            if value['weeknumber'] == week:
                if value.get("fact_budget") is not None and value.get("fact_impressions") is not None:
                    return_cell.value = value.get("fact_budget")*1000/value.get("fact_impressions")
                else:
                    return_cell.value = "N/A"
    if attrib == "fact_cpt":
        for value in placement_dict['postclick']:
            if value['weeknumber'] == week:
                if value.get("fact_budget") is not None and value.get("fact_reach") is not None:
                    return_cell.value = value.get("fact_budget")*1000/value.get("fact_reach")
                else:
                    return_cell.value = "N/A"
    if attrib == "fact_ctr":
        for value in placement_dict['postclick']:
            if value['weeknumber'] == week:
                if value.get("fact_clicks") is not None and value.get("fact_impressions") is not None:
                    return_cell.value = value.get("fact_clicks")/value.get("fact_impressions")
                else:
                    return_cell.value = "N/A"
    if attrib == "fact_cpc":
        for value in placement_dict['postclick']:
            if value['weeknumber'] == week:
                if value.get("fact_budget") is not None and value.get("fact_clicks") is not None:
                    return_cell.value = value.get("fact_budget")/value.get("fact_clicks")
                else:
                    return_cell.value = "N/A"
    if attrib == 'fact_vtr':
        for value in placement_dict['postclick']:
            if value['weeknumber'] == week:
                if value.get("fact_views") is not None and value.get("fact_impressions") is not None:
                    return_cell.value = value.get("fact_views")/value.get("fact_impressions")
                else:
                    return_cell.value = "N/A"
    if attrib == "fact_cpv":
        for value in placement_dict['postclick']:
            if value['weeknumber'] == week:
                if value.get("fact_budget") is not None and value.get("fact_views") is not None:
                    return_cell.value = value.get("fact_budget")/value.get("fact_views")
                else:
                    return_cell.value = "N/A"
    if attrib == 'period':
        return_cell.value = str(len(placement_dict.get('plan_weeks'))) + " weeks"
    if attrib == 'fact_impressions_adriver':
        if placement_dict.get('adriver_id') is not None:
            return_cell.value = get_adriver_value(placement_dict.get('adriver_id'),week)
        else:
            return_cell.value = "N/A"
    return return_cell


def get_total(attrib_cell, ws):
    if attrib_cell.value in ('plan_impressions', 'plan_reach', 'plan_clicks', 'plan_views', 'plan_budget', 'fact_impressions', 'fact_impressions_adriver', 'fact_reach', 'fact_clicks', 'fact_views', 'fact_budget'):
        return "=SUM(" + attrib_cell.column + "1:" + attrib_cell.column + str(ws.max_row-1) + ")"
    if attrib_cell.value == 'plan_cpm':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'plan_impressions':
                    plan_impressions_column = cell.column
                if cell.value == 'plan_budget':
                    plan_budget_column = cell.column
        return "=" + plan_budget_column + str(ws.max_row) + "*1000/" + plan_impressions_column + str(ws.max_row)
    if attrib_cell.value == 'fact_cpm':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'fact_impressions':
                    fact_impressions_column = cell.column
                if cell.value == 'fact_budget':
                    fact_budget_column = cell.column
        return "=" + fact_budget_column + str(ws.max_row) + "*1000/" + fact_impressions_column + str(ws.max_row)
    if attrib_cell.value == 'plan_cpt':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'plan_reach':
                    plan_reach_column = cell.column
                if cell.value == 'plan_budget':
                    plan_budget_column = cell.column
        return "=" + plan_budget_column + str(ws.max_row) + "*1000/" + plan_reach_column + str(ws.max_row)
    if attrib_cell.value == 'fact_cpt':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'fact_reach':
                    fact_reach_column = cell.column
                if cell.value == 'fact_budget':
                    fact_budget_column = cell.column
        return "=" + fact_budget_column + str(ws.max_row) + "*1000/" + fact_reach_column + str(ws.max_row)
    if attrib_cell.value == 'plan_ctr':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'plan_clicks':
                    plan_clicks_column = cell.column
                if cell.value == 'plan_impressions':
                    plan_impressions_column = cell.column
        return "=" + plan_clicks_column + str(ws.max_row) + "/" + plan_impressions_column + str(ws.max_row)
    if attrib_cell.value == 'fact_ctr':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'fact_clicks':
                    fact_clicks_column = cell.column
                if cell.value == 'fact_impressions':
                    fact_impressions_column = cell.column
        return "=" + fact_clicks_column + str(ws.max_row) + "/" + fact_impressions_column + str(ws.max_row)
    if attrib_cell.value == 'plan_cpc':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'plan_budget':
                    plan_budget_column = cell.column
                if cell.value == 'plan_clicks':
                    plan_clicks_column = cell.column
        return "=" + plan_budget_column + str(ws.max_row) + "/" + plan_clicks_column + str(ws.max_row)
    if attrib_cell.value == 'fact_cpc':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'fact_budget':
                    fact_budget_column = cell.column
                if cell.value == 'fact_clicks':
                    fact_clicks_column = cell.column
        return "=" + fact_budget_column + str(ws.max_row) + "/" + fact_clicks_column + str(ws.max_row)
    if attrib_cell.value == 'plan_cpv':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'plan_budget':
                    plan_budget_column = cell.column
                if cell.value == 'plan_views':
                    plan_views_column = cell.column
        return "=" + plan_budget_column + str(ws.max_row) + "/" + plan_views_column + str(ws.max_row)
    if attrib_cell.value == 'fact_cpv':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'fact_budget':
                    fact_budget_column = cell.column
                if cell.value == 'fact_views':
                    fact_views_column = cell.column
        return "=" + fact_budget_column + str(ws.max_row) + "/" + fact_views_column + str(ws.max_row)
    if attrib_cell.value == 'plan_vtr':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'plan_impressions':
                    plan_impressions_column = cell.column
                if cell.value == 'plan_views':
                    plan_views_column = cell.column
        return "=SUM(" + plan_views_column + "1:" + plan_views_column + str(ws.max_row-1) + ")/SUMIF(" + plan_views_column + "1:" + plan_views_column + str(ws.max_row-1) + ',">1",' +  plan_impressions_column + "1:" + plan_impressions_column + str(ws.max_row-1) + ")"
    if attrib_cell.value == 'fact_vtr':
        for row in ws.iter_rows(min_row=attrib_cell.row, max_row=attrib_cell.row):
            for cell in row:
                if cell.value == 'fact_impressions':
                    fact_impressions_column = cell.column
                if cell.value == 'fact_views':
                    fact_views_column = cell.column
        return "=SUM(" + fact_views_column + "1:" + fact_views_column + str(ws.max_row-1) + ")/SUMIF(" + fact_views_column + "1:" + fact_views_column + str(ws.max_row-1) + ',">1",' +  fact_impressions_column + "1:" + fact_impressions_column + str(ws.max_row-1) + ")"
    return ""



def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-o', '--operation', choices=['new', 'amnet', 'iprospect', 'report'], required=True)
    parser.add_argument('-f', '--file', required=False)
    parser.add_argument('-w', '--week', choices=[format(x,'d') for x in range(1,52)], required=False)
    res = parser.parse_args()
    if res.operation == 'new':
        parse_plan(res.file)
    if res.operation == 'amnet':
        parse_amnet(res.week)
    if res.operation == 'iprospect':
        parse_iprospect(res.week)
    if res.operation == 'report':
        create_report()

def get_ym_value(week):
        d = str(datetime.datetime.now().year) + '-W' + str(week)
        stop_date = datetime.datetime.strftime(datetime.datetime.strptime(d + '-1', "%Y-W%W-%w")+datetime.timedelta(days=6),"%Y-%m-%d")
        url = 'https://api-metrika.yandex.ru/stat/v1/data/bytime?metrics=ym:s:users&date1=2018-01-01&date2=' + str(stop_date)+ '&id=36885660&oauth_token=AQAAAAAeMXqkAAUhDB5LdCRAfU2rgaN6QJD9A8o'
        r = requests.get(url = url)
        print (r.content)

def updateDatabase(placement_id = None, field = None, value = None, week = None):
    PLACEMENT_ID = 'placement_id'
    LIST_OF_PLACEMENT_INFO_FIELDS = ['placement_info_category', 'placement_info_description', 'placement_info_format', 'placement_info_platform', 'placement_info_pricemodel', 'placement_info_stage']
    LIST_OF_PLACEMENT_PLAN_FIELDS = ['placement_plan_budget', 'placement_plan_clicks', 'placement_plan_impressions', 'placement_plan_reach', 'placement_plan_views', 'placement_plan_weeks']
    LIST_OF_PLACEMENT_STATS_FIELDS = ['placement_stats_adriverid', 'placement_stats_dcmid', 'placement_stats_tnscampaign','placement_stats_utmcampaign', 'placement_stats_ymcounter']
    client = MongoClient('localhost', 27017)
    db = client['VizeumHealth']
    collection = db['Placements']
    collection.update_one({PLACEMENT_ID:placement_id}, {"$set": {field:value}}, upsert=True)


if __name__ == '__main__':
    main()


