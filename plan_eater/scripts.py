from openpyxl import *
from openpyxl.cell import Cell
from openpyxl.formatting import Rule
from openpyxl.formatting.rule import *
from openpyxl.styles import Font, PatternFill, Border, Side, GradientFill
from openpyxl.styles.numbers import *
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.styles import Alignment
from openpyxl.worksheet.cell_range import CellRange
import re
import os
import datetime
import requests
import feedparser
from copy import copy


def create_new_json(plan):
    workbook = load_workbook(plan)
    mediaplan_sheet = workbook['MediaPlan']
    placement_dict = dict()
    weeknumber_set = set()
    for row in range(1,mediaplan_sheet.max_row):
        if re.search('\w\w_\w\w_\d\d\d', str(mediaplan_sheet.cell(row=row, column=1).value)):
            placement_dict.clear()
            weeknumber_set.clear()
            for column in range(2, mediaplan_sheet.max_column+1):

                if mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet), column=column).value is not None:
                    if re.search('\d{1,2}', str(mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet), column=column).value)):
                        if mediaplan_sheet.cell(row=row, column=column).value == 1:
                            weeknumber_set.add(int(mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet), column=column).value))
                        continue
                    placement_dict.update({mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet),
                                                                column=column).value: mediaplan_sheet.cell(
                        row=row, column=column).value})
            week_list = list()
            for weeknumber in sorted(weeknumber_set):
                week_list.append(
                    {'weeknumber': weeknumber, 'fact_budget': None, 'fact_impressions': None, 'fact_views': None,
                        'fact_clicks': None,
                        'fact_reach': None})
                placement_dict.update({'postclick': week_list})
            with open(str(os.path.dirname(plan)) + '\\' + str(
                    mediaplan_sheet.cell(row=row, column=1).value) + '.json', 'w') as outfile:
                json.dump(placement_dict, outfile)
    return


def update_json(plan):
    workbook = load_workbook(plan)
    mediaplan_sheet = workbook['MediaPlan']
    for row in range(1, mediaplan_sheet.max_row):
        if re.search('\w\w_\w\w_\d\d\d', str(mediaplan_sheet.cell(row=row, column=1).value)):
            with open(str(os.path.dirname(plan)) + '\\' + str(mediaplan_sheet.cell(row=row, column=1).value) + '.json','r') as infile:
                placement_dict = json.load(infile)
            old_week_list = list(value for value in placement_dict['postclick'] if value['weeknumber'] < datetime.datetime.today().isocalendar()[1])
            for column in range(2, mediaplan_sheet.max_column + 1):
                if mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet), column=column).value is not None:
                    if re.search('\d{1,2}',str(mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet), column=column).value)):
                        if (mediaplan_sheet.cell(row=row, column=column).value == 1) and (
                        datetime.datetime.today().isocalendar()[1] <= (mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet),column=column).value)):
                            if {'weeknumber': mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet),column=column).value, 'fact_budget': None, 'fact_impressions': None, 'fact_views': None,'fact_clicks': None, 'fact_reach': None} not in old_week_list:
                                old_week_list.append({'weeknumber': mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet),column=column).value, 'fact_budget': None, 'fact_impressions': None, 'fact_views': None,'fact_clicks': None, 'fact_reach': None})
                        continue
                    placement_dict.update({mediaplan_sheet.cell(row=get_fields_row(mediaplan_sheet),column=column).value:
                                               mediaplan_sheet.cell(row=row,column=column).value})
            with open(str(os.path.dirname(plan)) + '\\' + str(mediaplan_sheet.cell(row=row, column=1).value) + '.json','w') as outfile:
                json.dump(placement_dict, outfile)


def parse_amnet():
    week_number=datetime.datetime.today().isocalendar()[1]
    report = load_workbook(os.getcwd()+'\Reports\\Amnet\\Amnet'+str(week_number)+'.xlsx')
    week_report_sheet = report[str(week_number)]
    report_dict = dict()
    row = int()
    while row < week_report_sheet.max_row:
        row+=1
        if re.search('\w\w_\w\w_\d\d\d', str(week_report_sheet.cell(row=row, column=1).value)):
            placement_id = str(week_report_sheet.cell(row=row, column=1).value)
            while week_report_sheet.cell(row=row+1, column=1).value != None:
                report_dict.update({str(week_report_sheet.cell(row=row+1, column=1).value): week_report_sheet.cell(row=row+1, column=2).value})
                row+=1
            with open(str(os.getcwd()) + '\\MP\\' + placement_id + '.json', 'r+') as infile:
                placement_dict = json.load(infile)
                infile.seek(0)
                infile.truncate()
                week_list = placement_dict['postclick']
                for week_info in week_list:
                    if week_info['weeknumber'] == week_number:
                        for stat in week_info.keys():
                            week_info.update({stat:report_dict.get(stat)})
                            week_info.update({'weeknumber':week_number})
                placement_dict.update({'postclick': week_list})
                json.dump(placement_dict, infile)
                report_dict.clear()
    return


def create_report(week):
    #список недель для формирования необходимых листов
    report = load_workbook(os.getcwd()+'\Reports\\'+ 'Template.xlsx')
    week_set = set()
    for jsonplan in os.listdir(os.getcwd() + '\\MP\\'):
        if jsonplan.endswith('.json'):
            with open(str(os.getcwd()) + '\\MP\\' + jsonplan, 'r') as infile:
                placement_dict = json.load(infile)
                for value in placement_dict['postclick']:
                    if value['weeknumber'] <= datetime.datetime.today().isocalendar()[1]:
                        week_set.add(value['weeknumber'])

    for week in sorted(week_set):
        #начало переменных форматирования
        medium = Side(border_style="medium", color="000000")
        borders = Border (top=medium, left=medium, right=medium, bottom=medium)
        yellow_fill = GradientFill(stop=("ffff99", "ffff99"))
        black_fill = GradientFill(stop=("3b3b3b", "3b3b3b"))
        text_rotation = Alignment(textRotation=90)
        #конец переменных форматирования
        source = report.active
        target = report.copy_worksheet(source)
        fields_row=get_fields_row(target)
        first_stage_row= target.max_row+1

        for stage in ['Awareness','Consideration', 'Preference', 'Action', 'Loyalty']:
            if first_stage_row!=target.max_row+1:
                target.cell(row=first_stage_row, column = 2).alignment = text_rotation
                target.merge_cells(start_row=first_stage_row, end_row=target.max_row, start_column=2, end_column=2)
                style_merged_cells(target,first_stage_row,target.max_row,2,2,border=borders, fill = yellow_fill)
                first_stage_row=target.max_row+1
            for category in ['OLV', 'Programmatic', 'Social Media', 'SEA']:
                category_flag = True
                for jsonplan in os.listdir(os.getcwd() + '\\MP\\'):
                    if jsonplan.endswith('.json'):
                        with open(str(os.getcwd()) + '\\MP\\' + jsonplan, 'r') as infile:
                            placement_dict = json.load(infile)
                            if placement_dict["stage"]==stage and placement_dict["category"]==category:
                                last_column = 1
                                last_row = target.max_row+1
                                if category_flag:
                                    target.cell(row=last_row,column=last_column+1, value = stage)
                                    target.merge_cells(start_row=last_row, start_column=last_column+2,end_row=last_row,end_column=target.max_column-1)
                                    target.cell(row=last_row,column=last_column+2, value = category)
                                    style_merged_cells(target,last_row,last_row,last_column+2,target.max_column-1,border=borders,fill = yellow_fill)
                                    category_flag = False
                                    last_row+=1
                                while (target.cell(row=fields_row, column = last_column).value) != "end":
                                    cell = get_value_from_placement_dict(target.cell(row=fields_row, column = last_column).value,placement_dict)
                                    target.cell(row=last_row,column=last_column, value = cell.value)

                                    if target.cell(row=fields_row, column = last_column).value == "plan_impressions":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_impressions")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0_-₽'
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0_-'

                                    if target.cell(row=fields_row, column = last_column).value == "plan_reach":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_reach")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0_-'

                                    if target.cell(row=fields_row, column = last_column).value == "plan_clicks":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_clicks")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0_-'

                                    if target.cell(row=fields_row, column = last_column).value == "plan_views":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_views")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0_-'

                                    if target.cell(row=fields_row, column = last_column).value == "plan_budget":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_budget")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_impressions":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = value.get("fact_impressions"))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0_-'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_clicks":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = value.get("fact_clicks"))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0_-'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_budget":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = value.get("fact_budget"))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_reach":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = value.get("fact_reach"))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0_-'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_views":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = value.get("fact_views"))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0_-'

                                    if target.cell(row=fields_row, column = last_column).value == "plan_cpm":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_budget")*1000/placement_dict.get("plan_impressions")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "plan_cpt":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_budget")*1000/placement_dict.get("plan_reach")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "plan_ctr":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_clicks")/placement_dict.get("plan_impressions")))
                                        target.cell(row=last_row, column=last_column).number_format = FORMAT_PERCENTAGE_00

                                    if target.cell(row=fields_row, column = last_column).value == "plan_cpc":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_budget")/placement_dict.get("plan_clicks")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0.00_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "plan_vtr":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_views")/placement_dict.get("plan_impressions")))
                                        target.cell(row=last_row, column=last_column).number_format = FORMAT_PERCENTAGE_00

                                    if target.cell(row=fields_row, column = last_column).value == "plan_cpv":
                                        target.cell(row=last_row, column=last_column, value = (placement_dict.get("plan_budget")/placement_dict.get("plan_views")))
                                        target.cell(row=last_row, column=last_column).number_format = '#,##0.00_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_cpm":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = (value.get("fact_budget")*1000/value.get("fact_impressions")))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_cpt":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = (value.get("fact_budget")*1000/value.get("fact_reach")))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_ctr":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = (value.get("fact_clicks")/value.get("fact_impressions")))
                                                target.cell(row=last_row, column=last_column).number_format = FORMAT_PERCENTAGE_00

                                    if target.cell(row=fields_row, column = last_column).value == "fact_cpc":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = (value.get("fact_budget")/value.get("fact_clicks")))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0.00_-₽'

                                    if target.cell(row=fields_row, column = last_column).value == "fact_vtr":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = (value.get("fact_views")/value.get("fact_impressions")))
                                                target.cell(row=last_row, column=last_column).number_format = FORMAT_PERCENTAGE_00

                                    if target.cell(row=fields_row, column = last_column).value == "fact_cpv":
                                        for value in placement_dict['postclick']:
                                            if value['weeknumber'] == week:
                                                target.cell(row=last_row, column=last_column, value = (value.get("fact_budget")/value.get("fact_views")))
                                                target.cell(row=last_row, column=last_column).number_format = '#,##0.00_-₽'
                                    target.cell(row=last_row,column=last_column).border = copy(target.cell(row=fields_row, column = last_column).border)
                                    last_column+=1

        d = str(datetime.datetime.now().year) + '-W' + str(week)
        target.title = datetime.datetime.strftime(datetime.datetime.strptime(d + '-1', "%Y-W%W-%w"),"%d-%b-%Y") + " -- " + datetime.datetime.strftime(datetime.datetime.strptime(d + '-1', "%Y-W%W-%w")+datetime.timedelta(days=6),"%d-%b-%Y")
    report.save(os.getcwd()+'\Reports\\Client\\'+ 'Report_week' +str(datetime.datetime.today().isocalendar()[1])+'.xlsx')
    return


def style_merged_cells(ws, min_row, max_row, min_col, max_col, border=Border(), fill=None, font=None, alignment=None):
    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)
    first_cell = ws.cell(min_row,min_col)
    if alignment:
        ws.merge_cells(start_row=min_row,end_row=max_row,start_column=min_col,end_column=max_col)
        first_cell.alignment = alignment
    if font:
        first_cell.font = font
    for rows in ws.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row,max_row=max_row):
        for cell in rows:
            cell.border = cell.border + top
    for rows in ws.iter_rows(min_col=min_col, max_col=max_col,min_row=min_row, max_row=max_row):
        for cell in rows:
            cell.border = cell.border + bottom
    for columns in ws.iter_cols(min_col=min_col, max_col=max_col,min_row=min_row, max_row=max_row):
        for cell in columns:
            cell.border = cell.border + left
    for columns in ws.iter_cols(min_col=min_col, max_col=max_col,min_row=min_row, max_row=max_row):
        for cell in columns:
            cell.border = cell.border + right
            if fill:
                cell.fill=fill


def get_fields_row(worksheet):
    fields_row=1
    while worksheet.cell(row=fields_row,column=1).value != "fields":
        fields_row+=1
    return fields_row

def get_adriver_value():
        url = 'https://api.adriver.ru/stat/profiles/2737205/total'
        headers = {'content-type': 'application/atom+xml','X-Auth-UserID':'130529','X-Auth-Passwd':'%5C%A8%F2i%A7%8CD%98%12%F8I'}
        r = requests.get(url = url, headers=headers)
        feed = feedparser.parse(r.content)
        print (feed.entries)
        print (feed.entries[0]['adriver_exp'])


def get_value_from_placement_dict(attrib, placement_dict):
    wb = Workbook()
    ws = wb.active
    return_cell = Cell(ws)
    return_cell.value == ''
    if attrib == 'fields':
        return_cell.value == ''
        return return_cell
    if attrib == 'platform_site' or attrib == 'description' or attrib == 'format':
        return_cell.value = placement_dict.get(attrib)
        return return_cell
    return return_cell
    # if attrib == 'period':
    #     return str(len(placement_dict["postclick"])) + " weeks"
    # if re.search('\d{1,2}', str(attrib)):
    #     for x in (value for value in placement_dict['postclick'] if value['weeknumber'] == attrib):
    #         return 1

