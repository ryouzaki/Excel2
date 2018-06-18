from openpyxl import *
import re
import datetime
def create_new_json(plan):
    wb = load_workbook(plan)
    mediaplan_sheet = wb['MediaPlan']
    last_row = 1
    placement_dict = dict()
    weeknumber_set = set()
    week_list=list()
    while mediaplan_sheet.cell(row=last_row, column=1).value != "borders":
        if re.search('\w\w_\w\w_\d\d\d\d',str(mediaplan_sheet.cell(row=last_row, column=1).value)):
            print (last_row)
        last_row+=1
    i=1
    while mediaplan_sheet.cell(row = 4, column= i).value != "end":
        i+=1
        placement_dict.update({mediaplan_sheet.cell(row = 4, column= i).value:mediaplan_sheet.cell(row = 12, column= i).value})
        if ((i>9 and i<375) and mediaplan_sheet.cell(row = 12, column= i).value == 1):
            weeknumber_set.add((mediaplan_sheet.cell(row = 10, column= i).value).isocalendar()[1])
            print (weeknumber_set)
    for weeknumber in weeknumber_set:
        week_list.append({'weeknumber':weeknumber,'budget':None, 'impressions':None, 'views':None, 'clicks':None, 'reach':None})

    placement_dict.update({'postclick':week_list})


    with open('C:/Users/szaval01/Documents/24.05.2018/data2.json', 'w') as outfile:
        json.dump(placement_dict, outfile)
def update_json(plan):
    wb = load_workbook(plan)
    mediaplan_sheet = wb['MediaPlan']
    last_row = 1
    placement_dict = dict()
    weeknumber_set = set()
    week_list=list()
    while mediaplan_sheet.cell(row=last_row, column=1).value != "borders":
        if re.search('\w\w_\w\w_\d\d\d\d',str(mediaplan_sheet.cell(row=last_row, column=1).value)):
            print (last_row)
        last_row+=1
    i=1
    with open('C:/Users/szaval01/Documents/24.05.2018/data.json', 'r') as outfile:
        placement_dict = json.load(outfile)
        print (placement_dict['postclick'])
        week_list = placement_dict ['postclick']
    while mediaplan_sheet.cell(row = 4, column= i).value != "end":
        i+=1
        placement_dict.update({mediaplan_sheet.cell(row = 4, column= i).value:mediaplan_sheet.cell(row = 12, column= i).value})
        if ((i>9 and i<375) and
            (mediaplan_sheet.cell(row = 12, column= i).value == 1) and
                (datetime.datetime.today().isocalendar()[1] < (mediaplan_sheet.cell(row = 10, column= i).value).isocalendar()[1])):
            weeknumber_set.add((mediaplan_sheet.cell(row = 10, column= i).value).isocalendar()[1])
            print (weeknumber_set)
    for weeknumber in weeknumber_set:
        week_list.append({'weeknumber':weeknumber,'budget':None, 'impressions':None, 'views':None, 'clicks':None, 'reach':None})
    placement_dict.update({'postclick':week_list})
    with open('C:/Users/szaval01/Documents/24.05.2018/data2.json', 'w') as outfile:
        json.dump(placement_dict, outfile)
    print (week_list)
update_json('C:/Users/szaval01/Documents/Custom Office Templates/template.xltm')
xyu
